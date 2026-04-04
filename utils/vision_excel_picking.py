# utils/vision_excel_picking.py

from __future__ import annotations

import json
import re
from dataclasses import dataclass, asdict
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


DEFAULT_FILLRATE_PATH = "data/fillrate/latest/fillrate_latest.xlsx"
DEFAULT_SHEET_NAME = "Detalle de Entregas"


@dataclass
class ExcelPickingConfig:
    excel_path: str = DEFAULT_FILLRATE_PATH
    sheet_name: str = DEFAULT_SHEET_NAME
    header_row: int = 2          # 1-based
    first_data_row: int = 3      # 1-based
    include_empty_rows: bool = False


# =========================
# Normalización básica
# =========================
def _clean_text(txt: Any) -> str:
    s = "" if txt is None else str(txt)
    s = s.replace("\n", " ").replace("\r", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _norm_key(txt: Any) -> str:
    s = _clean_text(txt).lower()
    s = (
        s.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    s = re.sub(r"[^a-z0-9]+", "_", s).strip("_")
    return s


def _norm_code(txt: Any) -> str:
    s = _clean_text(txt).upper()
    s = s.replace(" ", "").replace(".", "").replace(",", "")
    return s

def _norm_estado_orden(txt: Any) -> str:
    """
    Normaliza Estado Orden preservando formato de 3 dígitos.
    Ej:
    40 -> "040"
    "040" -> "040"
    45 -> "045"
    """
    s = _clean_text(txt)
    if not s:
        return ""

    try:
        n = int(float(s))
        return f"{n:03d}"
    except Exception:
        s = re.sub(r"\D+", "", s)
        if not s:
            return ""
        return s.zfill(3)


def _to_int(val: Any) -> Optional[int]:
    if val in (None, ""):
        return None
    try:
        return int(round(float(val)))
    except Exception:
        return None


def _to_float(val: Any) -> Optional[float]:
    if val in (None, ""):
        return None
    try:
        return float(val)
    except Exception:
        return None


def _excel_datetime_to_date_str(val: Any) -> str:
    if val in (None, ""):
        return ""
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()

    s = _clean_text(val)
    return s


def _excel_datetime_to_time_str(val: Any) -> str:
    if val in (None, ""):
        return ""
    if isinstance(val, datetime):
        return val.strftime("%H:%M:%S")
    if isinstance(val, time):
        return val.strftime("%H:%M:%S")

    s = _clean_text(val)
    return s


def _normalize_excel_value(val: Any) -> Any:
    """
    Normaliza valores del Excel sin perder trazabilidad.
    - datetime/date -> ISO date
    - time/datetime-hora -> HH:MM:SS
    - float entero -> int
    - strings -> texto limpio
    - vacíos -> ""
    """
    if val is None:
        return ""

    if isinstance(val, datetime):
        # Si viene con hora 00:00:00, tratamos como fecha
        if val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.date().isoformat()
        return val.isoformat(sep=" ")

    if isinstance(val, date):
        return val.isoformat()

    if isinstance(val, time):
        return val.strftime("%H:%M:%S")

    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val

    s = _clean_text(val)
    if s == "":
        return ""

    try:
        f = float(s)
        if f.is_integer():
            return int(f)
        return f
    except Exception:
        return s


def _clean_record(rec: Dict[str, Any]) -> Dict[str, Any]:
    return {k: _normalize_excel_value(v) for k, v in rec.items()}


# =========================
# Encabezados esperados
# =========================
CANONICAL_HEADER_MAP = {
    "fecha_creacion": "fecha_creacion",
    "hora_creacion": "hora_creacion",
    "fecha_ultima_actualizacion": "fecha_ultima_actualizacion",
    "hora_ultima_actualizacion": "hora_ultima_actualizacion",
    "fecha_agenda": "fecha_agenda",
    "hora_agenda": "hora_agenda",
    "shipping": "shipping",
    "orden_compra": "orden_compra",
    "prioridad_de_despacho": "prioridad_despacho",
    "cliente": "cliente",
    "direccion": "direccion",
    "comuna": "comuna",
    "tipo_orden": "tipo_orden",
    "estado_orden": "estado_orden",
    "descripcion_estado": "descripcion_estado",
    "ruta": "ruta",
    "linea": "linea",
    "articulo": "articulo",
    "descripcion": "descripcion",
    "estado": "estado",
    "cant_original": "cant_original",
    "cant_trabajada": "cant_trabajada",
    "diferencia": "diferencia",
    "kits": "kits",
}


def _canonicalize_header(header: Any) -> str:
    norm = _norm_key(header)
    return CANONICAL_HEADER_MAP.get(norm, norm)

def filter_fillrate_for_packstructure(
    records: List[Dict[str, Any]],
    *,
    allowed_estado_orden: Optional[List[str]] = None,
) -> List[Dict[str, Any]]:
    """
    Filtra registros del FillRate que sí deben cruzarse con PackStructure.

    Regla actual:
    - Solo Estado Orden 040 y 045
      040 = Picking en proceso
      045 = Picking no iniciado
    """
    allowed_estado_orden = allowed_estado_orden or ["040", "045"]
    allowed_set = {_norm_estado_orden(x) for x in allowed_estado_orden if _norm_estado_orden(x)}

    out: List[Dict[str, Any]] = []

    for rec in records:
        estado_orden = _norm_estado_orden(rec.get("estado_orden"))
        articulo = _norm_code(rec.get("articulo"))

        if not articulo:
            continue

        if estado_orden in allowed_set:
            out.append(rec)

    return out


def extract_fillrate_skus_for_packstructure(
    records: List[Dict[str, Any]],
    *,
    allowed_estado_orden: Optional[List[str]] = None,
) -> List[str]:
    """
    Devuelve SKUs únicos del FillRate que deben cruzarse con PackStructure.
    """
    filtered = filter_fillrate_for_packstructure(
        records,
        allowed_estado_orden=allowed_estado_orden,
    )

    seen = set()
    skus: List[str] = []

    for rec in filtered:
        sku = _norm_code(rec.get("articulo"))
        if not sku or sku in seen:
            continue
        seen.add(sku)
        skus.append(sku)

    return skus


def build_fillrate_product_index(
    records: List[Dict[str, Any]],
    *,
    allowed_estado_orden: Optional[List[str]] = None,
) -> Dict[str, List[Dict[str, Any]]]:
    """
    Indexa FillRate por artículo (SKU), pero solo para los registros
    que deben participar del cruce con PackStructure.
    """
    filtered = filter_fillrate_for_packstructure(
        records,
        allowed_estado_orden=allowed_estado_orden,
    )

    idx: Dict[str, List[Dict[str, Any]]] = {}

    for rec in filtered:
        articulo = _norm_code(rec.get("articulo"))
        if not articulo:
            continue

        idx.setdefault(articulo, []).append({
            "row": rec.get("_row"),
            "shipping": _clean_text(rec.get("shipping")),
            "estado_orden": _clean_text(rec.get("estado_orden")),
            "descripcion_estado": _clean_text(rec.get("descripcion_estado")),
            "ruta": _clean_text(rec.get("ruta")),
            "linea": rec.get("linea"),
            "articulo": _clean_text(rec.get("articulo")),
            "descripcion": _clean_text(rec.get("descripcion")),
            "estado": _clean_text(rec.get("estado")),
            "cant_original": rec.get("cant_original"),
            "cant_trabajada": rec.get("cant_trabajada"),
            "diferencia": rec.get("diferencia"),
            "kits": rec.get("kits"),
            "cliente": _clean_text(rec.get("cliente")),
            "orden_compra": _clean_text(rec.get("orden_compra")),
            "excel_full": _clean_record(rec.get("excel_full", {})),
        })

    return idx

# =========================
# Funciones de filtrado
# =========================

def match_fillrate_with_packstructure(
    fillrate_result: Dict[str, Any],
    packstructure_result: Dict[str, Any],
    *,
    allowed_estado_orden: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """
    Cruza FillRate con PackStructure usando:
    FillRate.articulo == PackStructure.sku

    Solo considera registros FillRate con Estado Orden 040 o 045
    (o los que se indiquen en allowed_estado_orden).
    """
    fillrate_records = fillrate_result.get("records", []) or []
    pack_records = packstructure_result.get("records", []) or []

    fillrate_filtered = filter_fillrate_for_packstructure(
        fillrate_records,
        allowed_estado_orden=allowed_estado_orden,
    )

    fillrate_idx = build_fillrate_product_index(
        fillrate_records,
        allowed_estado_orden=allowed_estado_orden,
    )

    # import local para evitar acoplamiento circular duro si lo prefieres
    from utils.vision_excel_packStructure import (
        filter_packstructure_by_skus,
        build_packstructure_index,
    )

    target_skus = list(fillrate_idx.keys())
    pack_filtered = filter_packstructure_by_skus(
        pack_records,
        skus=target_skus,
    )
    pack_idx = build_packstructure_index(pack_filtered)

    matched_products: List[Dict[str, Any]] = []

    for sku in target_skus:
        fillrate_matches = fillrate_idx.get(sku, [])
        pack_matches = pack_idx.get(sku, [])

        if not pack_matches:
            matched_products.append({
                "sku_fillrate": sku,
                "match_status": "not_found_in_packstructure",
                "fillrate_match_count": len(fillrate_matches),
                "packstructure_match_count": 0,
                "fillrate_matches": fillrate_matches,
                "packstructure": None,
                "packstructure_matches": [],
            })
            continue

        first_pack = pack_matches[0]

        matched_products.append({
            "sku_fillrate": sku,
            "sku_packstructure": first_pack.get("sku"),
            "descripcion_fillrate": fillrate_matches[0].get("descripcion") if fillrate_matches else "",
            "descripcion_packstructure": first_pack.get("descripcion"),
            "match_status": "matched",
            "fillrate_match_count": len(fillrate_matches),
            "packstructure_match_count": len(pack_matches),

            # EANs clave
            "ean_ea": first_pack.get("ean"),
            "ean_in": first_pack.get("ean_in"),
            "ean_cs": first_pack.get("ean_cs"),

            # cantidades / niveles logísticos
            "qty_ea": first_pack.get("qty_ea"),
            "qty_inn": first_pack.get("qty_inn"),
            "qty_cs": first_pack.get("qty_cs"),
            "qty_pal": first_pack.get("qty_pal"),

            # dimensiones / pesos
            "largo_ea": first_pack.get("largo_ea"),
            "ancho_ea": first_pack.get("ancho_ea"),
            "alto_ea": first_pack.get("alto_ea"),
            "peso_ea": first_pack.get("peso_ea"),

            "largo_inn": first_pack.get("largo_inn"),
            "ancho_inn": first_pack.get("ancho_inn"),
            "alto_inn": first_pack.get("alto_inn"),
            "peso_inn": first_pack.get("peso_inn"),

            "largo_cs": first_pack.get("largo_cs"),
            "ancho_cs": first_pack.get("ancho_cs"),
            "alto_cs": first_pack.get("alto_cs"),
            "peso_cs": first_pack.get("peso_cs"),

            "largo_pal": first_pack.get("largo_pal"),
            "ancho_pal": first_pack.get("ancho_pal"),
            "alto_pal": first_pack.get("alto_pal"),
            "peso_pal": first_pack.get("peso_pal"),

            # trazabilidad
            "fillrate_matches": fillrate_matches,
            "packstructure": first_pack,
            "packstructure_matches": pack_matches,
        })

    return {
        "status": "success",
        "filter_applied": {
            "allowed_estado_orden": allowed_estado_orden or ["040", "045"],
            "fillrate_rows_total": len(fillrate_records),
            "fillrate_rows_filtered": len(fillrate_filtered),
            "unique_skus_filtered": len(target_skus),
            "packstructure_rows_total": len(pack_records),
            "packstructure_rows_filtered": len(pack_filtered),
        },
        "products": matched_products,
        "counts": {
            "matched_products": sum(1 for x in matched_products if x["match_status"] == "matched"),
            "not_found_products": sum(1 for x in matched_products if x["match_status"] == "not_found_in_packstructure"),
        },
    }

# =========================
# Lectura principal
# =========================
def _read_sheet_rows(excel_path: str, sheet_name: str) -> List[List[Any]]:
    wb = load_workbook(excel_path, data_only=True, read_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"No existe la hoja '{sheet_name}'. Hojas disponibles: {wb.sheetnames}"
        )

    ws = wb[sheet_name]
    rows: List[List[Any]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    return rows


def load_excel_picking(
    excel_path: Optional[str] = None,
    *,
    cfg: Optional[ExcelPickingConfig] = None,
) -> Dict[str, Any]:
    cfg = cfg or ExcelPickingConfig()
    final_excel_path = excel_path or cfg.excel_path

    excel_file = Path(final_excel_path)
    if not excel_file.exists():
        return {
            "status": "error",
            "error": "excel_not_found",
            "excel_path": str(excel_file),
        }

    try:
        rows = _read_sheet_rows(str(excel_file), cfg.sheet_name)
    except Exception as exc:
        return {
            "status": "error",
            "error": "sheet_read_error",
            "excel_path": str(excel_file),
            "sheet_name": cfg.sheet_name,
            "message": str(exc),
        }

    if not rows:
        return {
            "status": "error",
            "error": "empty_sheet",
            "excel_path": str(excel_file),
            "sheet_name": cfg.sheet_name,
        }

    if cfg.header_row < 1 or cfg.header_row > len(rows):
        return {
            "status": "error",
            "error": "invalid_header_row",
            "excel_path": str(excel_file),
            "sheet_name": cfg.sheet_name,
            "header_row": cfg.header_row,
        }

    raw_headers = rows[cfg.header_row - 1]
    original_headers = [_clean_text(h) for h in raw_headers]
    normalized_headers = [_canonicalize_header(h) for h in original_headers]

    records: List[Dict[str, Any]] = []

    for excel_row_num in range(cfg.first_data_row, len(rows) + 1):
        row_vals = rows[excel_row_num - 1]
        rec_raw: Dict[str, Any] = {}
        rec_norm: Dict[str, Any] = {}

        has_content = False

        for idx, key in enumerate(normalized_headers):
            value = row_vals[idx] if idx < len(row_vals) else ""

            if value not in ("", None):
                has_content = True

            rec_raw[original_headers[idx] if idx < len(original_headers) else f"col_{idx+1}"] = _normalize_excel_value(value)

            # tratamiento especial por tipo de campo
            if key in {
                "fecha_creacion",
                "fecha_ultima_actualizacion",
                "fecha_agenda",
            }:
                rec_norm[key] = _excel_datetime_to_date_str(value)
            elif key in {
                "hora_creacion",
                "hora_ultima_actualizacion",
                "hora_agenda",
            }:
                rec_norm[key] = _excel_datetime_to_time_str(value)
            elif key in {"cant_original", "cant_trabajada", "diferencia", "linea"}:
                rec_norm[key] = _to_int(value)
            else:
                rec_norm[key] = _normalize_excel_value(value)

        if not has_content and not cfg.include_empty_rows:
            continue

        rec_norm["_row"] = excel_row_num
        rec_norm["_source"] = {
            "sheet_name": cfg.sheet_name,
            "excel_path": str(excel_file),
        }
        rec_norm["excel_full"] = rec_raw

        records.append(rec_norm)

    key_fields = {
        "shipping": sum(1 for r in records if _clean_text(r.get("shipping"))),
        "estado_orden": sum(1 for r in records if _clean_text(r.get("estado_orden"))),
        "descripcion_estado": sum(1 for r in records if _clean_text(r.get("descripcion_estado"))),
        "ruta": sum(1 for r in records if _clean_text(r.get("ruta"))),
        "cant_original": sum(1 for r in records if r.get("cant_original") is not None),
    }

    return {
        "status": "success",
        "excel_path": str(excel_file),
        "sheet_name": cfg.sheet_name,
        "rows_loaded": len(records),
        "header_row": cfg.header_row,
        "first_data_row": cfg.first_data_row,
        "headers": {
            "original": original_headers,
            "normalized": normalized_headers,
        },
        "important_fields_presence": key_fields,
        "records": records,
        "config": asdict(cfg),
    }

# =========================
# Funcion de Summarize
# =========================

def summarize_fillrate_for_packstructure(
    records: List[Dict[str, Any]],
    *,
    allowed_estado_orden: Optional[List[str]] = None,
) -> Dict[str, Any]:
    filtered = filter_fillrate_for_packstructure(
        records,
        allowed_estado_orden=allowed_estado_orden,
    )

    skus = extract_fillrate_skus_for_packstructure(
        records,
        allowed_estado_orden=allowed_estado_orden,
    )

    shipping_vals = sorted({
        _clean_text(r.get("shipping"))
        for r in filtered
        if _clean_text(r.get("shipping"))
    })
    rutas = sorted({
        _clean_text(r.get("ruta"))
        for r in filtered
        if _clean_text(r.get("ruta"))
    })

    total_cant_original = sum((r.get("cant_original") or 0) for r in filtered)

    return {
        "rows_filtered": len(filtered),
        "unique_skus_filtered": len(skus),
        "unique_shipping_count": len(shipping_vals),
        "unique_ruta_count": len(rutas),
        "total_cant_original_filtered": total_cant_original,
        "sample_shipping": shipping_vals[:10],
        "sample_rutas": rutas[:10],
        "sample_skus": skus[:20],
    }

# =========================
# Helpers útiles para el siguiente paso
# =========================
def filter_excel_records(
    records: List[Dict[str, Any]],
    *,
    shipping: Optional[str] = None,
    ruta: Optional[str] = None,
    estado_orden: Optional[str] = None,
    descripcion_estado: Optional[str] = None,
    articulo: Optional[str] = None,
) -> List[Dict[str, Any]]:
    shipping_n = _norm_code(shipping)
    ruta_n = _norm_code(ruta)
    estado_orden_n = _norm_code(estado_orden)
    descripcion_estado_n = _norm_code(descripcion_estado)
    articulo_n = _norm_code(articulo)

    out: List[Dict[str, Any]] = []

    for rec in records:
        ok = True

        if shipping_n and _norm_code(rec.get("shipping")) != shipping_n:
            ok = False
        if ruta_n and _norm_code(rec.get("ruta")) != ruta_n:
            ok = False
        if estado_orden_n and _norm_code(rec.get("estado_orden")) != estado_orden_n:
            ok = False
        if descripcion_estado_n and _norm_code(rec.get("descripcion_estado")) != descripcion_estado_n:
            ok = False
        if articulo_n and _norm_code(rec.get("articulo")) != articulo_n:
            ok = False

        if ok:
            out.append(rec)

    return out


def build_excel_product_index(records: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
    """
    Indexa por artículo y conserva TODAS las filas asociadas a ese artículo.
    Útil para el cruce posterior con PackStructure.
    """
    idx: Dict[str, List[Dict[str, Any]]] = {}

    for rec in records:
        articulo = _norm_code(rec.get("articulo"))
        if not articulo:
            continue

        idx.setdefault(articulo, []).append({
            "row": rec.get("_row"),
            "shipping": _clean_text(rec.get("shipping")),
            "estado_orden": _clean_text(rec.get("estado_orden")),
            "descripcion_estado": _clean_text(rec.get("descripcion_estado")),
            "ruta": _clean_text(rec.get("ruta")),
            "linea": rec.get("linea"),
            "articulo": _clean_text(rec.get("articulo")),
            "descripcion": _clean_text(rec.get("descripcion")),
            "estado": _clean_text(rec.get("estado")),
            "cant_original": rec.get("cant_original"),
            "cant_trabajada": rec.get("cant_trabajada"),
            "diferencia": rec.get("diferencia"),
            "kits": rec.get("kits"),
            "cliente": _clean_text(rec.get("cliente")),
            "orden_compra": _clean_text(rec.get("orden_compra")),
            "excel_full": _clean_record(rec.get("excel_full", {})),
        })

    return idx


def summarize_excel_records(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    unique_shipping = sorted({
        _clean_text(r.get("shipping"))
        for r in records
        if _clean_text(r.get("shipping"))
    })
    unique_rutas = sorted({
        _clean_text(r.get("ruta"))
        for r in records
        if _clean_text(r.get("ruta"))
    })
    unique_estado_orden = sorted({
        _clean_text(r.get("estado_orden"))
        for r in records
        if _clean_text(r.get("estado_orden"))
    })
    unique_descripcion_estado = sorted({
        _clean_text(r.get("descripcion_estado"))
        for r in records
        if _clean_text(r.get("descripcion_estado"))
    })

    total_cant_original = sum(
        r.get("cant_original", 0) or 0
        for r in records
    )
    total_cant_trabajada = sum(
        r.get("cant_trabajada", 0) or 0
        for r in records
    )
    total_diferencia = sum(
        r.get("diferencia", 0) or 0
        for r in records
    )

    return {
        "rows": len(records),
        "unique_shipping_count": len(unique_shipping),
        "unique_ruta_count": len(unique_rutas),
        "unique_estado_orden_count": len(unique_estado_orden),
        "unique_descripcion_estado_count": len(unique_descripcion_estado),
        "total_cant_original": total_cant_original,
        "total_cant_trabajada": total_cant_trabajada,
        "total_diferencia": total_diferencia,
        "sample_shipping": unique_shipping[:10],
        "sample_rutas": unique_rutas[:10],
        "sample_estado_orden": unique_estado_orden[:10],
        "sample_descripcion_estado": unique_descripcion_estado[:10],
    }


def save_json(data: Dict[str, Any], out_path: str) -> str:
    p = Path(out_path)
    p.parent.mkdir(parents=True, exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return str(p)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Carga y procesa el FillRate Excel más reciente.")
    parser.add_argument(
        "excel_path",
        nargs="?",
        default=DEFAULT_FILLRATE_PATH,
        help="Ruta al Excel FillRate. Default: data/fillrate/latest/fillrate_latest.xlsx",
    )
    parser.add_argument(
        "--sheet-name",
        default=DEFAULT_SHEET_NAME,
        help="Nombre de hoja a procesar",
    )
    parser.add_argument(
        "--out-json",
        default=None,
        help="Ruta opcional para guardar salida JSON",
    )
    parser.add_argument(
        "--summary-only",
        action="store_true",
        help="Si se activa, imprime solo el resumen y no todos los records",
    )
    args = parser.parse_args()

    cfg = ExcelPickingConfig(
        excel_path=args.excel_path,
        sheet_name=args.sheet_name,
    )

    res = load_excel_picking(args.excel_path, cfg=cfg)

    if res.get("status") == "success":
        summary = summarize_excel_records(res.get("records", []))
        res["summary"] = summary

    if args.summary_only and res.get("status") == "success":
        printable = {
            "status": res["status"],
            "excel_path": res["excel_path"],
            "sheet_name": res["sheet_name"],
            "rows_loaded": res["rows_loaded"],
            "headers": res["headers"],
            "important_fields_presence": res["important_fields_presence"],
            "summary": res.get("summary", {}),
            "config": res["config"],
        }
    else:
        printable = res

    if args.out_json:
        save_json(printable, args.out_json)

    print(json.dumps(printable, ensure_ascii=False, indent=2))
    
# salida completa;
# python -m utils.vision_excel_picking

# guardar .json
# python -m utils.vision_excel_picking --out-json results/fillrate_processed.json