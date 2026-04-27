# utils/report.py
from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


REPORT_SHEETS = {
    "summary": "Resumen",
    "products": "DetalleProductos",
    "barcodes": "Barcodes",
    "captures": "Capturas",
}


SUMMARY_HEADERS = [
    "processed_at_local",
    "session_id",
    "mode_app",
    "shipping",
    "shipping_normalized",
    "ruta",
    "target_sku",
    "closure_status",
    "session_status",
    "expected_products",
    "matched_products",
    "partial_products",
    "missing_products",
    "excess_products",
    "expected_units",
    "observed_units",
    "difference_units",
    "detected_barcodes_count",
    "detected_unique_barcodes_count",
    "unknown_barcodes_count",
    "contaminated_barcodes_count",
    "captures_processed_count",
    "summary_json",
    "readout_json",
    "closure_json",
]


PRODUCT_HEADERS = [
    "processed_at_local",
    "session_id",
    "shipping",
    "ruta",
    "closure_status",
    "sku",
    "descripcion",
    "expected_units",
    "observed_units",
    "difference_units",
    "product_status",
    "ean_ea",
    "qty_ea",
    "ean_in",
    "qty_in",
    "ean_cs",
    "qty_cs",
    "barcode_hits",
    "closure_json",
]


BARCODE_HEADERS = [
    "processed_at_local",
    "session_id",
    "shipping",
    "ruta",
    "closure_status",
    "barcode",
    "sku",
    "descripcion",
    "level",
    "units_per_barcode",
    "count",
    "classification",
    "source",
    "closure_json",
]


CAPTURE_HEADERS = [
    "processed_at_local",
    "session_id",
    "shipping",
    "ruta",
    "closure_status",
    "capture_processed_at",
    "event_dir",
    "readout_json",
    "detected_barcodes_json",
    "new_barcodes_count",
    "new_barcodes_unique_count",
    "new_barcodes",
    "closure_json",
]


def _read_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        payload = json.load(f)
    if not isinstance(payload, dict):
        raise ValueError(f"El JSON no contiene un objeto raíz válido: {path}")
    return payload


def _safe_list(value: Any) -> List[Any]:
    return value if isinstance(value, list) else []


def _safe_dict(value: Any) -> Dict[str, Any]:
    return value if isinstance(value, dict) else {}


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _norm_barcode(value: Any) -> str:
    return "".join(ch for ch in _norm_text(value) if ch.isdigit())


def _norm_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except Exception:
        return default


def _join_values(values: Any) -> str:
    if isinstance(values, list):
        return ", ".join(_norm_text(x) for x in values if _norm_text(x))
    if isinstance(values, dict):
        return json.dumps(values, ensure_ascii=False)
    return _norm_text(values)


def _status_from_payload(payload: Dict[str, Any]) -> str:
    return (
        _norm_text(payload.get("effective_closure_status"))
        or _norm_text(_safe_dict(payload.get("closure_result")).get("closure_status"))
        or _norm_text(_safe_dict(payload.get("frontend_summary")).get("closure_status"))
        or _norm_text(_safe_dict(payload.get("session")).get("last_closure_status"))
    )


def _get_session_id(payload: Dict[str, Any]) -> str:
    session = _safe_dict(payload.get("session"))
    return _norm_text(session.get("session_id")) or f"session_unknown_{_norm_text(payload.get('processed_at_local'))}"


def _get_shipping(payload: Dict[str, Any]) -> str:
    session = _safe_dict(payload.get("session"))
    resolution = _safe_dict(payload.get("target_shipping_resolution"))
    closure_result = _safe_dict(payload.get("closure_result"))
    event_context = _safe_dict(closure_result.get("event_context"))

    return (
        _norm_text(session.get("target_shipping"))
        or _norm_text(resolution.get("target_shipping"))
        or _norm_text(event_context.get("target_shipping"))
    )


def _get_shipping_normalized(payload: Dict[str, Any]) -> str:
    closure_result = _safe_dict(payload.get("closure_result"))
    event_context = _safe_dict(closure_result.get("event_context"))
    return _norm_text(event_context.get("target_shipping_normalized"))


def _get_route(payload: Dict[str, Any]) -> str:
    session = _safe_dict(payload.get("session"))
    resolution = _safe_dict(payload.get("target_shipping_resolution"))
    frontend = _safe_dict(payload.get("frontend_summary"))
    closure_result = _safe_dict(payload.get("closure_result"))

    return (
        _norm_text(session.get("target_ruta"))
        or _norm_text(resolution.get("target_ruta"))
        or _norm_text(frontend.get("route"))
        or _norm_text(closure_result.get("route"))
    )


def _get_target_sku(payload: Dict[str, Any]) -> str:
    session = _safe_dict(payload.get("session"))
    resolution = _safe_dict(payload.get("target_shipping_resolution"))
    return _norm_text(session.get("target_sku")) or _norm_text(resolution.get("target_sku"))


def _closure_json_path(path: Path) -> str:
    return str(path)


def _build_summary_row(payload: Dict[str, Any], closure_json_path: Path) -> Dict[str, Any]:
    frontend = _safe_dict(payload.get("frontend_summary"))
    frontend_totals = _safe_dict(frontend.get("totals"))

    closure_result = _safe_dict(payload.get("closure_result"))
    closure_counts = _safe_dict(closure_result.get("counts"))
    closure_totals = _safe_dict(closure_result.get("totals"))

    session = _safe_dict(payload.get("session"))

    detected_barcodes = _safe_list(closure_result.get("detected_barcodes"))
    unique_barcodes = sorted({_norm_barcode(x) for x in detected_barcodes if _norm_barcode(x)})

    captures = _safe_list(session.get("captures_processed"))

    return {
        "processed_at_local": payload.get("processed_at_local"),
        "session_id": _get_session_id(payload),
        "mode_app": payload.get("mode_app"),
        "shipping": _get_shipping(payload),
        "shipping_normalized": _get_shipping_normalized(payload),
        "ruta": _get_route(payload),
        "target_sku": _get_target_sku(payload),
        "closure_status": _status_from_payload(payload),
        "session_status": session.get("session_status"),
        "expected_products": closure_counts.get("expected_products", frontend_totals.get("products_expected", 0)),
        "matched_products": closure_counts.get("matched_products", frontend_totals.get("products_matched", 0)),
        "partial_products": closure_counts.get("partial_products", frontend_totals.get("products_partial", 0)),
        "missing_products": closure_counts.get("missing_products", frontend_totals.get("products_missing", 0)),
        "excess_products": closure_counts.get("excess_products", 0),
        "expected_units": closure_totals.get("expected_units", session.get("target_shipping_expected_units", 0)),
        "observed_units": closure_totals.get("observed_units", session.get("target_shipping_observed_units", 0)),
        "difference_units": closure_totals.get("difference_units", 0),
        "detected_barcodes_count": len(detected_barcodes) or frontend_totals.get("detected_barcodes_count", 0),
        "detected_unique_barcodes_count": len(unique_barcodes),
        "unknown_barcodes_count": closure_counts.get("unknown_barcodes", 0),
        "contaminated_barcodes_count": closure_counts.get("contaminated_barcodes", 0),
        "captures_processed_count": len(captures),
        "summary_json": payload.get("summary_json"),
        "readout_json": payload.get("readout_json"),
        "closure_json": _closure_json_path(closure_json_path),
    }


def _build_product_rows(payload: Dict[str, Any], closure_json_path: Path) -> List[Dict[str, Any]]:
    closure_result = _safe_dict(payload.get("closure_result"))
    products = _safe_list(closure_result.get("products"))
    session_id = _get_session_id(payload)
    shipping = _get_shipping(payload)
    route = _get_route(payload)
    status = _status_from_payload(payload)
    processed_at = payload.get("processed_at_local")

    rows: List[Dict[str, Any]] = []

    for product in products:
        if not isinstance(product, dict):
            continue

        levels = _safe_dict(product.get("expected_pack_levels"))
        ea = _safe_dict(levels.get("EA"))
        inn = _safe_dict(levels.get("IN"))
        cs = _safe_dict(levels.get("CS"))

        observed_barcodes = _safe_list(product.get("observed_barcodes"))
        barcode_counter = Counter(
            _norm_barcode(x.get("barcode"))
            for x in observed_barcodes
            if isinstance(x, dict) and _norm_barcode(x.get("barcode"))
        )

        rows.append(
            {
                "processed_at_local": processed_at,
                "session_id": session_id,
                "shipping": shipping,
                "ruta": route,
                "closure_status": status,
                "sku": product.get("sku"),
                "descripcion": product.get("descripcion"),
                "expected_units": product.get("expected_units"),
                "observed_units": product.get("observed_units"),
                "difference_units": product.get("difference_units"),
                "product_status": product.get("status"),
                "ean_ea": ea.get("ean"),
                "qty_ea": ea.get("units"),
                "ean_in": inn.get("ean"),
                "qty_in": inn.get("units"),
                "ean_cs": cs.get("ean"),
                "qty_cs": cs.get("units"),
                "barcode_hits": "; ".join(f"{k}: {v}" for k, v in barcode_counter.items()),
                "closure_json": _closure_json_path(closure_json_path),
            }
        )

    return rows


def _build_barcode_rows(payload: Dict[str, Any], closure_json_path: Path) -> List[Dict[str, Any]]:
    closure_result = _safe_dict(payload.get("closure_result"))
    session_id = _get_session_id(payload)
    shipping = _get_shipping(payload)
    route = _get_route(payload)
    status = _status_from_payload(payload)
    processed_at = payload.get("processed_at_local")

    rows: List[Dict[str, Any]] = []

    for product in _safe_list(closure_result.get("products")):
        if not isinstance(product, dict):
            continue
        for item in _safe_list(product.get("observed_barcodes")):
            if not isinstance(item, dict):
                continue
            rows.append(
                {
                    "processed_at_local": processed_at,
                    "session_id": session_id,
                    "shipping": shipping,
                    "ruta": route,
                    "closure_status": status,
                    "barcode": item.get("barcode"),
                    "sku": product.get("sku"),
                    "descripcion": product.get("descripcion"),
                    "level": item.get("level"),
                    "units_per_barcode": item.get("units_per_barcode"),
                    "count": 1,
                    "classification": "valid",
                    "source": item.get("source"),
                    "closure_json": _closure_json_path(closure_json_path),
                }
            )

    for item in _safe_list(closure_result.get("contaminated_barcodes")):
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "processed_at_local": processed_at,
                "session_id": session_id,
                "shipping": shipping,
                "ruta": route,
                "closure_status": status,
                "barcode": item.get("barcode"),
                "sku": item.get("sku"),
                "descripcion": item.get("descripcion"),
                "level": item.get("level"),
                "units_per_barcode": item.get("units_per_barcode"),
                "count": 1,
                "classification": "contaminated",
                "source": item.get("reason"),
                "closure_json": _closure_json_path(closure_json_path),
            }
        )

    for barcode in _safe_list(closure_result.get("unknown_barcodes")):
        rows.append(
            {
                "processed_at_local": processed_at,
                "session_id": session_id,
                "shipping": shipping,
                "ruta": route,
                "closure_status": status,
                "barcode": barcode,
                "sku": "",
                "descripcion": "",
                "level": "",
                "units_per_barcode": 0,
                "count": 1,
                "classification": "unknown",
                "source": "unknown_barcodes",
                "closure_json": _closure_json_path(closure_json_path),
            }
        )

    return rows


def _build_capture_rows(payload: Dict[str, Any], closure_json_path: Path) -> List[Dict[str, Any]]:
    session = _safe_dict(payload.get("session"))
    captures = _safe_list(session.get("captures_processed"))

    session_id = _get_session_id(payload)
    shipping = _get_shipping(payload)
    route = _get_route(payload)
    status = _status_from_payload(payload)
    processed_at = payload.get("processed_at_local")

    rows: List[Dict[str, Any]] = []

    for capture in captures:
        if not isinstance(capture, dict):
            continue

        rows.append(
            {
                "processed_at_local": processed_at,
                "session_id": session_id,
                "shipping": shipping,
                "ruta": route,
                "closure_status": status,
                "capture_processed_at": capture.get("processed_at_local"),
                "event_dir": capture.get("event_dir"),
                "readout_json": capture.get("readout_json"),
                "detected_barcodes_json": capture.get("detected_barcodes_json"),
                "new_barcodes_count": capture.get("new_barcodes_count"),
                "new_barcodes_unique_count": capture.get("new_barcodes_unique_count"),
                "new_barcodes": _join_values(capture.get("new_barcodes")),
                "closure_json": _closure_json_path(closure_json_path),
            }
        )

    return rows


def _ensure_workbook(report_path: Path) -> Workbook:
    if report_path.exists():
        return load_workbook(report_path)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    for sheet_name in REPORT_SHEETS.values():
        wb.create_sheet(sheet_name)

    return wb


def _ensure_headers(ws, headers: List[str]) -> None:
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(headers)
        return

    existing = [ws.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]
    if existing != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)


def _delete_existing_session_rows(ws, session_id: str, closure_json: str) -> None:
    if ws.max_row <= 1:
        return

    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    try:
        session_col = headers.index("session_id") + 1
    except ValueError:
        return

    closure_col = headers.index("closure_json") + 1 if "closure_json" in headers else None

    rows_to_delete: List[int] = []
    for row in range(2, ws.max_row + 1):
        row_session = _norm_text(ws.cell(row=row, column=session_col).value)
        row_closure = _norm_text(ws.cell(row=row, column=closure_col).value) if closure_col else ""
        if row_session == session_id or (closure_json and row_closure == closure_json):
            rows_to_delete.append(row)

    for row in reversed(rows_to_delete):
        ws.delete_rows(row)


def _append_dict_rows(ws, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    for row in rows:
        ws.append([row.get(header) for header in headers])


def _style_sheet(ws, headers: List[str]) -> None:
    header_fill = PatternFill("solid", fgColor="DCEBFA")
    header_font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9E2EC")
    border = Border(bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        width = 14

        if header in {"descripcion", "summary_json", "readout_json", "closure_json", "event_dir", "new_barcodes"}:
            width = 42
        elif header in {"session_id", "processed_at_local", "capture_processed_at"}:
            width = 24
        elif header in {"closure_status", "product_status", "classification"}:
            width = 18
        elif header in {"barcode", "ean_ea", "ean_in", "ean_cs"}:
            width = 20
        elif header in {"shipping", "shipping_normalized", "ruta", "target_sku", "sku"}:
            width = 18
        elif header.endswith("_count") or header.endswith("_units") or header in {"count", "qty_ea", "qty_in", "qty_cs"}:
            width = 14

        ws.column_dimensions[col_letter].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18


def _apply_status_styles(ws, headers: List[str]) -> None:
    if "closure_status" not in headers:
        return

    status_col = headers.index("closure_status") + 1
    fills = {
        "complete_match": PatternFill("solid", fgColor="DFF3E3"),
        "target_shipping_complete": PatternFill("solid", fgColor="DFF3E3"),
        "partial_match": PatternFill("solid", fgColor="FFF3CD"),
        "contaminated": PatternFill("solid", fgColor="F8D7DA"),
        "no_detection": PatternFill("solid", fgColor="E5E7EB"),
    }

    for row in range(2, ws.max_row + 1):
        value = _norm_text(ws.cell(row=row, column=status_col).value)
        fill = fills.get(value)
        if not fill:
            continue
        ws.cell(row=row, column=status_col).fill = fill


def append_closure_report(
    closure_json_path: Path,
    report_xlsx: Path = Path("data/reports/picking_closure_report.xlsx"),
) -> Path:
    closure_json_path = Path(closure_json_path)
    report_xlsx = Path(report_xlsx)

    payload = _read_json(closure_json_path)

    session_id = _get_session_id(payload)
    closure_json_str = _closure_json_path(closure_json_path)

    summary_row = _build_summary_row(payload, closure_json_path)
    product_rows = _build_product_rows(payload, closure_json_path)
    barcode_rows = _build_barcode_rows(payload, closure_json_path)
    capture_rows = _build_capture_rows(payload, closure_json_path)

    report_xlsx.parent.mkdir(parents=True, exist_ok=True)

    wb = _ensure_workbook(report_xlsx)

    sheet_specs = [
        (REPORT_SHEETS["summary"], SUMMARY_HEADERS, [summary_row]),
        (REPORT_SHEETS["products"], PRODUCT_HEADERS, product_rows),
        (REPORT_SHEETS["barcodes"], BARCODE_HEADERS, barcode_rows),
        (REPORT_SHEETS["captures"], CAPTURE_HEADERS, capture_rows),
    ]

    for sheet_name, headers, rows in sheet_specs:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
        _ensure_headers(ws, headers)
        _delete_existing_session_rows(ws, session_id=session_id, closure_json=closure_json_str)
        _append_dict_rows(ws, headers, rows)
        _style_sheet(ws, headers)
        _apply_status_styles(ws, headers)

    wb.save(report_xlsx)
    return report_xlsx


def build_report_from_folder(
    closure_dir: Path,
    report_xlsx: Path = Path("data/reports/picking_closure_report.xlsx"),
    pattern: str = "*.json",
) -> Path:
    closure_dir = Path(closure_dir)
    files = sorted(closure_dir.glob(pattern))

    if not files:
        raise FileNotFoundError(f"No encontré JSONs en {closure_dir} con patrón {pattern}")

    for path in files:
        try:
            append_closure_report(path, report_xlsx=report_xlsx)
        except Exception as exc:
            print(f"[WARN] No pude agregar {path}: {exc}")

    return report_xlsx


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera reporte Excel acumulativo desde resultados de closure.")
    parser.add_argument("closure_json", nargs="?", help="Ruta a closure_output JSON.")
    parser.add_argument("--report-xlsx", default="data/reports/picking_closure_report.xlsx")
    parser.add_argument("--from-folder", default=None, help="Carpeta con JSONs de closure para consolidar.")
    parser.add_argument("--pattern", default="*.json")

    args = parser.parse_args()

    report_path = Path(args.report_xlsx)

    if args.from_folder:
        out = build_report_from_folder(
            Path(args.from_folder),
            report_xlsx=report_path,
            pattern=args.pattern,
        )
    else:
        if not args.closure_json:
            raise SystemExit("Debes indicar closure_json o usar --from-folder")
        out = append_closure_report(
            Path(args.closure_json),
            report_xlsx=report_path,
        )

    print(f"[OK] Reporte generado/actualizado en: {out}")


if __name__ == "__main__":
    main()